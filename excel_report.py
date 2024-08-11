import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def create_excel(author_data, main_table_data, file_name="report.xlsx"):
    # Создаем новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Заголовок и данные автора
    ws["A1"] = "Author"
    ws["B1"] = author_data['author']
    ws["A2"] = "Creation Date"
    ws["B2"] = author_data['creation_date']
    ws["A3"] = "Project Name"
    ws["B3"] = author_data['project_name']
    ws["A4"] = "Power (Wp)"
    ws["B4"] = author_data['power_wp']

    # Переходим на следующую строку после введения данных автора
    start_row = 6

    # Форматирование заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Заголовки основной таблицы
    for col_num, header in enumerate(main_table_data[0], 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Заполнение основной таблицы
    for row_num, row_data in enumerate(main_table_data[1:], start=start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Форматирование ширины столбцов
    for col in range(1, len(main_table_data[0]) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Сохранение файла
    wb.save(file_name)
    return file_name
